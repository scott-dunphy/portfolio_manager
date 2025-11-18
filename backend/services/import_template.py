from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


PROPERTY_HEADERS = [
    "Property_ID",
    "Property_Name",
    "Property_Type",
    "Address",
    "City",
    "State",
    "Zip_Code",
    "Purchase_Date",
    "Purchase_Price",
    "Market_Value_Start",
    "Initial_NOI",
    "NOI_Growth_Rate",
    "Capex_Percent_of_NOI",
    "Use_Manual_NOI_Capex",
    "Ownership_Percent",
    "Exit_Date",
    "Exit_Cap_Rate",
    "Building_Size",
    "Valuation_Method",
]

LOAN_HEADERS = [
    "Loan_ID",
    "Loan_Name",
    "Property_ID",
    "Principal_Amount",
    "Rate_Type",
    "Interest_Rate",
    "SOFR_Spread",
    "Origination_Date",
    "Maturity_Date",
    "Payment_Frequency",
    "Loan_Type",
    "Amortization_Period_Months",
    "IO_Period_Months",
    "Origination_Fee",
    "Exit_Fee",
]

LOAN_CASH_FLOW_HEADERS = [
    "Loan_ID",
    "Payment_Date",
    "Interest_Amount",
    "Principal_Amount",
]

MANUAL_HEADERS = [
    "Property_ID",
    "Year",
    "Frequency",  # annual or monthly
    "Month (1-12, required if monthly)",
    "NOI",
    "Capex",
]


def _auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)


def build_import_template():
    wb = Workbook()

    instructions = wb.active
    instructions.title = "Instructions"
    instructions["A1"] = "Portfolio Import Template"
    instructions["A1"].font = Font(size=16, bold=True)

    instructions.append([])
    instructions.append(["How to use:"])
    instructions.append(["1. Fill out the Properties sheet (Property_ID is required and must be unique per portfolio)."])
    instructions.append(["2. (Optional) Add loans tied to each property in the Loans sheet using Property_ID to link."])
    instructions.append(["3. (Optional) Provide manual Annual NOI/Capex overrides per property/year in the Manual_NOI_Capex sheet."])
    instructions.append(["4. (Optional) Provide actual loan cash flows for specific months in the Loan_Cash_Flows sheet. Only enter the months that should override calculated principal or interest."])
    instructions.append(["5. Save the workbook and upload it from the Upload tab, selecting the target portfolio."])
    instructions.append(["Notes: Dates can be entered in any Excel date format. Percentages should be decimals (0.1 = 10%). Market_Value_Start represents the property value at the portfolio analysis start date."])

    instructions.column_dimensions["A"].width = 120

    ws_properties = wb.create_sheet("Properties")
    ws_properties.append(PROPERTY_HEADERS)
    ws_properties.append(
        [
            "PROP-1001",
            "Dunphy Towers",
            "Office",
            "123 Main St",
            "New York",
            "NY",
            "10001",
            "2024-01-01",
            200000000,
            210000000,
            12000000,
            0.03,
            0.05,
            "No",
            1.0,
            "2029-06-30",
            0.055,
            500000,
            "growth",
        ]
    )
    _auto_width(ws_properties)

    ws_loans = wb.create_sheet("Loans")
    ws_loans.append(LOAN_HEADERS)
    ws_loans.append(
        [
            "LOAN-1",
            "Senior Loan",
            "PROP-1001",
            120000000,
            "fixed",
            0.05,
            "",
            "2024-01-01",
            "2029-01-01",
            "monthly",
            "Senior",
            360,
            0,
            0,
            0,
        ]
    )
    _auto_width(ws_loans)

    ws_manual = wb.create_sheet("Manual_NOI_Capex")
    ws_manual.append(MANUAL_HEADERS)
    ws_manual.append(["PROP-1001", 2024, "annual", "", 12000000, 600000])
    ws_manual.append(["PROP-1001", 2024, "monthly", 1, 1000000, 50000])
    _auto_width(ws_manual)

    ws_loan_cf = wb.create_sheet("Loan_Cash_Flows")
    ws_loan_cf.append(LOAN_CASH_FLOW_HEADERS)
    ws_loan_cf.append(["LOAN-1", "2024-03-31", 500000, 250000])
    ws_loan_cf.append(["LOAN-1", "2024-04-30", 480000, ""])
    _auto_width(ws_loan_cf)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
