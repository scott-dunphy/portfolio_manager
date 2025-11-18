from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


def export_property_type_exposure_to_excel(exposure_data: dict, portfolio_name: str = "Portfolio") -> BytesIO:
    """
    Export property type exposure data to Excel format.

    Args:
        exposure_data: Dictionary with 'dates', 'property_types', 'data', and 'transactions'
        portfolio_name: Name of the portfolio for the filename

    Returns:
        BytesIO stream containing the Excel file
    """
    wb = Workbook()

    # Remove default sheet and create new ones
    wb.remove(wb.active)

    # Create Exposure Summary sheet
    ws_exposure = wb.create_sheet("Quarterly Exposure")

    # Headers
    headers = ["Quarter End Date"] + exposure_data.get("property_types", [])
    ws_exposure.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws_exposure[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    data = exposure_data.get("data", [])
    dates = exposure_data.get("dates", [])
    property_types = exposure_data.get("property_types", [])

    for idx, row_data in enumerate(data):
        date_str = dates[idx] if idx < len(dates) else ""
        # Format date
        if date_str:
            date_obj = datetime.fromisoformat(date_str)
            date_str = date_obj.strftime("%b %d, %Y")

        row = [date_str]
        for ptype in property_types:
            type_data = row_data.get(ptype, {})
            if isinstance(type_data, dict):
                percentage = type_data.get("percentage", 0)
            else:
                percentage = type_data or 0
            row.append(f"{percentage:.1f}%")

        ws_exposure.append(row)

    # Auto-width columns
    _auto_width(ws_exposure)

    # Create Market Values sheet
    ws_values = wb.create_sheet("Market Values")

    headers = ["Quarter End Date"] + property_types
    ws_values.append(headers)

    # Style header row
    for cell in ws_values[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows with market values
    for idx, row_data in enumerate(data):
        date_str = dates[idx] if idx < len(dates) else ""
        # Format date
        if date_str:
            date_obj = datetime.fromisoformat(date_str)
            date_str = date_obj.strftime("%b %d, %Y")

        row = [date_str]
        for ptype in property_types:
            type_data = row_data.get(ptype, {})
            if isinstance(type_data, dict):
                market_value = type_data.get("market_value", 0)
            else:
                market_value = 0
            row.append(market_value)

        ws_values.append(row)

    # Format currency columns
    for row in ws_values.iter_rows(min_row=2, min_col=2, max_col=len(property_types) + 1):
        for cell in row:
            cell.number_format = '$#,##0'

    _auto_width(ws_values)

    # Create Transactions sheet if available
    transactions = exposure_data.get("transactions", [])
    if transactions:
        ws_trans = wb.create_sheet("Acquisitions & Dispositions")

        trans_headers = ["Transaction Date", "Property Name", "Property Type", "Transaction Type", "Transaction Price"]
        ws_trans.append(trans_headers)

        # Style header row
        for cell in ws_trans[1]:
            cell.fill = header_fill
            cell.font = header_font

        for trans in transactions:
            date_obj = datetime.fromisoformat(trans.get("transaction_date", ""))
            date_str = date_obj.strftime("%b %d, %Y")

            trans_type = trans.get("transaction_type", "")
            trans_type_display = "Acquisition" if trans_type == "acquisition" else "Disposition"

            row = [
                date_str,
                trans.get("property_name", ""),
                trans.get("property_type", ""),
                trans_type_display,
                trans.get("transaction_price", 0)
            ]
            ws_trans.append(row)

        # Format price column as currency
        for row in ws_trans.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '$#,##0'

        _auto_width(ws_trans)

    # Save to stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)
