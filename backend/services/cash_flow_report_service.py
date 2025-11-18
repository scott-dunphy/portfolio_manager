from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from calendar import monthrange
from io import BytesIO
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from database import db
from models import CashFlow, Loan, Portfolio, Property
from services.property_valuation_service import calculate_property_valuation


def build_cash_flow_report(portfolio_id: int) -> BytesIO:
    portfolio = Portfolio.query.get_or_404(portfolio_id)

    loans: List[Loan] = Loan.query.filter_by(portfolio_id=portfolio_id).all()
    loan_map: Dict[int, Loan] = {loan.id: loan for loan in loans}

    cash_flows: List[CashFlow] = (
        CashFlow.query.filter_by(portfolio_id=portfolio_id)
        .order_by(CashFlow.date)
        .all()
    )

    properties: List[Property] = (
        Property.query.filter_by(portfolio_id=portfolio_id).all()
    )
    property_map: Dict[int, Property] = {prop.id: prop for prop in properties}
    ownership_lookup = _prepare_ownership_lookup(properties)
    valuation_lookup, valuation_series = _prepare_property_valuations(properties)
    property_outstanding, property_outstanding_ownership = _compute_property_outstanding(
        properties, loans, cash_flows, ownership_lookup
    )

    type_set = set()
    aggregate_by_date: Dict[date, dict] = {}
    property_rows: List[dict] = []
    property_rows_ownership: List[dict] = []
    capex_by_property_date: Dict[tuple, float] = defaultdict(float)

    for cf in cash_flows:
        cf_date = cf.date
        cf_type = cf.cash_flow_type or 'Uncategorized'
        type_set.add(cf_type)

        base_amount = cf.amount or 0.0
        property_id = cf.property_id

        if property_id and property_id in property_map:
            ownership_percent = _ownership_percent(
                ownership_lookup[property_id],
                cf_date,
                default=property_map[property_id].ownership_percent or 1.0
            )
        else:
            ownership_percent = 1.0

        ownership_amount = base_amount * ownership_percent if property_id else base_amount

        date_entry = aggregate_by_date.setdefault(
            cf_date,
            {
                'date': cf_date,
                'total': 0.0,
                'total_ownership': 0.0,
                'type_totals': defaultdict(float),
                'type_totals_ownership': defaultdict(float),
            },
        )
        date_entry['total'] += base_amount
        date_entry['total_ownership'] += ownership_amount
        date_entry['type_totals'][cf_type] += base_amount
        date_entry['type_totals_ownership'][cf_type] += ownership_amount
        if property_id:
            outstanding_value = property_outstanding.get(property_id, {}).get(cf_date)
            if outstanding_value is not None:
                date_entry['loan_outstanding_total'] = (date_entry.get('loan_outstanding_total') or 0.0) + outstanding_value
            outstanding_value_own = property_outstanding_ownership.get(property_id, {}).get(cf_date)
            if outstanding_value_own is not None:
                date_entry['loan_outstanding_total_ownership'] = (date_entry.get('loan_outstanding_total_ownership') or 0.0) + outstanding_value_own

        property_label, loan_label = _resolve_labels(property_map, loan_map, cf)

        if property_id and cf_type == 'property_capex':
            capex_by_property_date[(property_id, cf_date)] += base_amount

        property_rows.append(
            {
                'property_id': property_id,
                'ownership_percent': 1.0,
                'property': property_label,
                'date': cf_date,
                'loan': loan_label,
                'type': cf_type,
                'amount': base_amount,
                'description': cf.description or '',
                'outstanding_debt': property_outstanding.get(property_id, {}).get(cf_date),
            }
        )
        property_rows_ownership.append(
            {
                'property_id': property_id,
                'ownership_percent': ownership_percent,
                'property': property_label,
                'date': cf_date,
                'loan': loan_label,
                'type': cf_type,
                'amount': ownership_amount,
                'description': cf.description or '',
                'outstanding_debt': property_outstanding_ownership.get(property_id, {}).get(cf_date),
            }
        )

    _annotate_property_rows(
        property_rows,
        valuation_lookup,
        capex_by_property_date,
        ownership_lookup,
        property_map,
        scale_with_ownership=False,
    )
    _annotate_property_rows(
        property_rows_ownership,
        valuation_lookup,
        capex_by_property_date,
        ownership_lookup,
        property_map,
        scale_with_ownership=True,
    )

    _augment_aggregate_with_valuations(
        aggregate_by_date,
        valuation_series,
        capex_by_property_date,
        property_map,
        ownership_lookup,
    )

    workbook = Workbook()
    type_headers = sorted(type_set)

    _build_aggregate_sheet(
        workbook.active,
        title='Fund Aggregate 100%',
        aggregate_by_date=aggregate_by_date,
        type_headers=type_headers,
        beginning_cash=portfolio.beginning_cash or 0.0,
        ownership=False,
    )

    sheet2 = workbook.create_sheet('Fund Aggregate Ownership')
    _build_aggregate_sheet(
        sheet2,
        title='Fund Aggregate Ownership',
        aggregate_by_date=aggregate_by_date,
        type_headers=type_headers,
        beginning_cash=portfolio.beginning_cash or 0.0,
        ownership=True,
    )

    sheet3 = workbook.create_sheet('Property Detail 100%')
    _build_property_sheet(sheet3, property_rows)

    sheet4 = workbook.create_sheet('Property Detail Ownership')
    _build_property_sheet(sheet4, property_rows_ownership)

    # Add Property Summary sheet
    sheet5 = workbook.create_sheet('Property Summary')
    _build_property_summary_sheet(sheet5, properties, valuation_lookup)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _build_aggregate_sheet(ws, title, aggregate_by_date, type_headers, beginning_cash, ownership=False):
    ws.title = title
    header = [
        'Date',
        'Beginning Cash',
        'Total',
        'Ending Cash',
        'Market Value (Current)',
        'Market Value (Prior)',
        'Appreciation',
        'Outstanding Debt',
    ] + type_headers
    ws.append(header)
    ws.row_dimensions[1].font = Font(bold=True)

    running_cash_100 = beginning_cash
    running_cash_own = beginning_cash
    sorted_dates = sorted(aggregate_by_date.keys())

    for cf_date in sorted_dates:
        entry = aggregate_by_date[cf_date]
        total = entry['total']
        total_ownership = entry['total_ownership']

        total_used = total_ownership if ownership else total
        running_begin = running_cash_own if ownership else running_cash_100
        running_end = running_begin + total_used

        market_current = entry.get('market_value_current_ownership') if ownership else entry.get('market_value_current')
        market_prior = entry.get('market_value_prior_ownership') if ownership else entry.get('market_value_prior')
        appreciation_total = entry.get('appreciation_total_ownership') if ownership else entry.get('appreciation_total')

        loan_outstanding = entry.get('loan_outstanding_total_ownership') if ownership else entry.get('loan_outstanding_total')
        row = [
            cf_date.isoformat(),
            round(running_begin),
            round(total_used),
            round(running_end),
            round(market_current or 0.0),
            round(market_prior or 0.0),
            round(appreciation_total or 0.0),
            round(loan_outstanding or 0.0),
        ]

        type_totals = entry['type_totals_ownership'] if ownership else entry['type_totals']
        for type_name in type_headers:
            row.append(round(type_totals.get(type_name, 0.0)))

        ws.append(row)

        if ownership:
            running_cash_own = running_end
        else:
            running_cash_100 = running_end

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 40)


def _build_property_sheet(ws, rows):
    ws.append([
        'Property',
        'Date',
        'Loan',
        'Type',
        'Amount',
        'Current Market Value',
        'Prior Market Value',
        'Appreciation',
        'Forward NOI (12m)',
        'Cap Rate',
        'Outstanding Debt',
        'Description'
    ])
    ws.row_dimensions[1].font = Font(bold=True)

    for row in sorted(rows, key=lambda item: (item['property'], item['date'])):
        ws.append([
            row['property'],
            row['date'].isoformat(),
            row['loan'],
            row['type'],
            round(row['amount'] or 0.0),
            round(row.get('market_value') or 0.0),
            round(row.get('market_value_prior') or 0.0),
            round(row.get('appreciation') or 0.0),
            round(row.get('forward_noi_12m') or 0.0),
            row.get('cap_rate'),
            round(row.get('outstanding_debt') or 0.0),
            row['description'],
        ])

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 50)


def _annotate_property_rows(
    rows,
    valuation_lookup,
    capex_by_property_date,
    ownership_lookup,
    property_map,
    scale_with_ownership=False,
):
    for row in rows:
        property_id = row.get('property_id')
        date_value = row.get('date')
        if not property_id or not date_value:
            continue

        property_lookup = valuation_lookup.get(property_id)
        if not property_lookup:
            continue

        valuation_entry = property_lookup.get(date_value)
        if not valuation_entry:
            continue

        current = valuation_entry.get('market_value')
        prior = valuation_entry.get('market_value_prior')
        forward_noi = valuation_entry.get('forward_noi_12m')
        cap_rate = valuation_entry.get('cap_rate')
        capex = -(capex_by_property_date.get((property_id, date_value), 0.0) or 0.0)
        appreciation = (
            current - prior - capex
            if current is not None and prior is not None
            else None
        )

        scale = row.get('ownership_percent') if scale_with_ownership else 1.0

        row['market_value'] = current * scale if current is not None else None
        row['market_value_prior'] = prior * scale if prior is not None else None
        row['forward_noi_12m'] = forward_noi * scale if forward_noi is not None else None
        row['cap_rate'] = cap_rate
        row['appreciation'] = appreciation * scale if appreciation is not None else None


def _augment_aggregate_with_valuations(
    aggregate_by_date,
    valuation_series,
    capex_by_property_date,
    property_map,
    ownership_lookup,
):
    for property_id, entries in valuation_series.items():
        property_obj = property_map.get(property_id)
        if not property_obj:
            continue
        exit_cutoff = _month_end(property_obj.exit_date) if property_obj.exit_date else None
        for record in entries:
            entry_date = record.get('date')
            if not entry_date:
                continue
            if exit_cutoff and entry_date > exit_cutoff:
                break
            date_entry = _ensure_date_entry(aggregate_by_date, entry_date)
            current = record.get('market_value')
            prior = record.get('market_value_prior')
            if current is None or prior is None:
                continue
            capex = -(capex_by_property_date.get((property_id, entry_date), 0.0) or 0.0)
            appreciation = current - prior - capex

            date_entry['market_value_current'] = (date_entry.get('market_value_current') or 0.0) + current
            date_entry['market_value_prior'] = (date_entry.get('market_value_prior') or 0.0) + prior
            date_entry['appreciation_total'] = (date_entry.get('appreciation_total') or 0.0) + appreciation

            ownership_percent = _ownership_percent(
                ownership_lookup.get(property_id, []),
                entry_date,
                default=property_obj.ownership_percent or 1.0
            )
            date_entry['market_value_current_ownership'] = (
                (date_entry.get('market_value_current_ownership') or 0.0) + current * ownership_percent
            )
            date_entry['market_value_prior_ownership'] = (
                (date_entry.get('market_value_prior_ownership') or 0.0) + prior * ownership_percent
            )
            date_entry['appreciation_total_ownership'] = (
                (date_entry.get('appreciation_total_ownership') or 0.0) + appreciation * ownership_percent
            )


def _ensure_date_entry(aggregate_by_date, target_date):
    return aggregate_by_date.setdefault(
        target_date,
        {
            'date': target_date,
            'total': 0.0,
            'total_ownership': 0.0,
            'type_totals': defaultdict(float),
            'type_totals_ownership': defaultdict(float),
        },
    )


def _prepare_ownership_lookup(properties: List[Property]) -> Dict[int, List[Tuple[date, float]]]:
    lookup = {}
    for property_obj in properties:
        events = sorted(property_obj.ownership_events, key=lambda event: event.event_date or date.min)
        lookup[property_obj.id] = [(event.event_date, event.ownership_percent) for event in events]
    return lookup


def _ownership_percent(events: List[Tuple[date, float]], target_date: date, default: float = 1.0) -> float:
    percent = default or 1.0
    for event_date, event_percent in events:
        if not event_date:
            continue
        if event_date <= target_date:
            percent = event_percent
        else:
            break
    return percent


def _prepare_property_valuations(properties: List[Property]):
    lookup = {}
    series = {}
    for property_obj in properties:
        valuation = calculate_property_valuation(property_obj)
        monthly_values = valuation.get('monthly_market_values') or []
        entries = []
        entry_lookup = {}
        prev_value = property_obj.market_value_start or 0.0
        exit_cutoff = _month_end(property_obj.exit_date) if property_obj.exit_date else None
        for item in monthly_values:
            date_str = item.get('date')
            if not date_str:
                continue
            try:
                entry_date = datetime.fromisoformat(date_str).date()
            except ValueError:
                continue
            if exit_cutoff and entry_date > exit_cutoff:
                break
            current = item.get('market_value')
            record = {
                'date': entry_date,
                'market_value': current,
                'market_value_prior': prev_value,
                'forward_noi_12m': item.get('forward_noi_12m'),
                'cap_rate': item.get('cap_rate'),
            }
            entries.append(record)
            entry_lookup[entry_date] = record
            if current is not None:
                prev_value = current
        lookup[property_obj.id] = entry_lookup
        series[property_obj.id] = entries
    return lookup, series


def _month_end(value: Optional[date]) -> Optional[date]:
    if value is None:
        return None
    last_day = monthrange(value.year, value.month)[1]
    return value.replace(day=last_day)


def _compute_property_outstanding(
    properties: List[Property],
    loans: List[Loan],
    cash_flows: List[CashFlow],
    ownership_lookup: Dict[int, List[Tuple[date, float]]],
) -> Tuple[Dict[int, Dict[date, float]], Dict[int, Dict[date, float]]]:
    property_map = {prop.id: prop for prop in properties}
    date_points = sorted({cf.date for cf in cash_flows if cf.date})
    outstanding = defaultdict(dict)
    outstanding_ownership = defaultdict(dict)
    if not date_points:
        return outstanding, outstanding_ownership

    flows_by_loan = defaultdict(list)
    for cf in cash_flows:
        if cf.loan_id and cf.date:
            flows_by_loan[cf.loan_id].append(cf)

    for flow_list in flows_by_loan.values():
        flow_list.sort(key=lambda cf: (cf.date, cf.id or 0))

    for loan in loans:
        prop_id = loan.property_id
        if not prop_id:
            continue
        flows = flows_by_loan.get(loan.id, [])
        pointer = 0
        balance = 0.0
        for current_date in date_points:
            while pointer < len(flows) and flows[pointer].date <= current_date:
                cf = flows[pointer]
                amount = cf.amount or 0.0
                flow_type = (cf.cash_flow_type or '').lower()
                if flow_type == 'loan_funding':
                    balance += amount
                elif flow_type == 'loan_principal':
                    balance += amount
                pointer += 1
            outstanding[prop_id][current_date] = outstanding[prop_id].get(current_date, 0.0) + balance
            percent = _ownership_percent(
                ownership_lookup.get(prop_id, []),
                current_date,
                default=property_map.get(prop_id).ownership_percent if property_map.get(prop_id) else 1.0
            )
            outstanding_ownership[prop_id][current_date] = outstanding_ownership[prop_id].get(current_date, 0.0) + balance * percent

    return outstanding, outstanding_ownership


def _resolve_labels(property_map, loan_map, cf):
    if cf.property_id and cf.property_id in property_map:
        property_obj = property_map[cf.property_id]
        property_label = property_obj.property_name or property_obj.property_id or f"Property #{property_obj.id}"
    else:
        property_label = "Unassigned"

    if cf.loan_id and cf.loan_id in loan_map:
        loan_obj = loan_map[cf.loan_id]
        loan_label = loan_obj.loan_name or loan_obj.loan_id or f"Loan #{loan_obj.id}"
    else:
        loan_label = '—'

    if property_label == "Unassigned" and loan_label != '—':
        property_label = f"Unassigned ({loan_label})"

    return property_label, loan_label


def _build_property_summary_sheet(ws, properties: List[Property], valuation_lookup: Dict):
    """Build a summary sheet with underlying property data."""
    ws.title = 'Property Summary'

    headers = [
        'Property ID',
        'Property Name',
        'Property Type',
        'Address',
        'City',
        'State',
        'Zip Code',
        'Purchase Date',
        'Purchase Price',
        'Market Value Start',
        'Initial NOI',
        'NOI Growth Rate',
        'Exit Date',
        'Exit Cap Rate',
        'Ownership %',
        'Building Size',
        'Valuation Method'
    ]
    ws.append(headers)
    ws.row_dimensions[1].font = Font(bold=True)

    for prop in properties:
        row = [
            prop.property_id or '',
            prop.property_name or '',
            prop.property_type or '',
            prop.address or '',
            prop.city or '',
            prop.state or '',
            prop.zip_code or '',
            prop.purchase_date.strftime('%Y-%m-%d') if prop.purchase_date else '',
            prop.purchase_price or '',
            prop.market_value_start or '',
            prop.initial_noi or '',
            prop.noi_growth_rate or '',
            prop.exit_date.strftime('%Y-%m-%d') if prop.exit_date else '',
            prop.exit_cap_rate or '',
            (prop.ownership_percent or 1.0) * 100,  # Convert to percentage
            prop.building_size or '',
            prop.valuation_method or ''
        ]
        ws.append(row)

    # Format columns
    ws.column_dimensions['H'].number_format = 'YYYY-MM-DD'
    ws.column_dimensions['I'].number_format = '$#,##0'
    ws.column_dimensions['J'].number_format = '$#,##0'
    ws.column_dimensions['K'].number_format = '$#,##0'
    ws.column_dimensions['L'].number_format = '0.0%'
    ws.column_dimensions['N'].number_format = 'YYYY-MM-DD'
    ws.column_dimensions['O'].number_format = '0.0%'
    ws.column_dimensions['P'].number_format = '#,##0'

    # Auto-width columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 40)
